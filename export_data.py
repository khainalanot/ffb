#!/usr/bin/env python3
"""
Re-export player projections from the Excel file into data/rankings.json.

Pulls every position (QB/RB/WR/TE/DST) from its projection sheet, and merges
in Jake's QB color tags + spreadsheet comments from the "Jakes Ranks" sheet.

Run this any time the Excel file is updated:
    python3 export_data.py

Then commit + push the updated data/rankings.json. Ryan's own edits (tag/rank
changes and comments made on the site) live in the database and are layered on
top of this base data at load time, so re-exporting does NOT wipe them.
"""
import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

SCRIPT_DIR = Path(__file__).resolve().parent
XLSX_PATH = SCRIPT_DIR.parent / "2026-FFB-Projections-0805-1 (1).xlsx"
OUTPUT_PATH = SCRIPT_DIR / "data" / "rankings.json"
SEED_PATH = SCRIPT_DIR / "seed_comments.sql"

# On the "Jakes Ranks" sheet there are four side-by-side ranking tables; each
# hand-written comment is anchored somewhere in its table's column block and
# belongs to the player in that table's "Player" column on the same row.
JAKES_TABLE_PLAYER_COL = [
    (1, 14, 2),    # QB block  -> player in col B (2)
    (16, 28, 17),  # RB block  -> player in col Q (17)
    (30, 41, 31),  # WR block  -> player in col AE (31)
    (43, 49, 44),  # TE block  -> player in col AR (44)
]
TS_RE = re.compile(r"\((\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})\)")

# Per-position column layout (1-indexed) and the stat fields to surface.
# Each entry: sheet name, column map, list of (stat_key, label, column).
POSITIONS = {
    "QB": {
        "sheet": "QB",
        "player": 2, "team": 3, "bye": 4, "fps": 13, "auction": 15,
        "stats": [
            ("Pass Yds", 7), ("Pass TD", 8), ("INT", 9),
            ("Rush Yds", 11), ("Rush TD", 12),
        ],
    },
    "RB": {
        "sheet": "RB",
        "player": 2, "team": 3, "bye": 4, "fps": 14, "auction": None,
        "stats": [
            ("Rush Yds", 6), ("Rush TD", 7), ("Rec", 9),
            ("Rec Yds", 10), ("Rec TD", 11), ("PPR", 14),
        ],
    },
    "WR": {
        "sheet": "WR",
        "player": 2, "team": 3, "bye": 4, "fps": 13, "auction": 15,
        "stats": [
            ("Rec", 8), ("Rec Yds", 9), ("Rec TD", 10),
            ("Rush Yds", 5), ("PPR", 13),
        ],
    },
    "TE": {
        "sheet": "TE",
        "player": 2, "team": 3, "bye": 4, "fps": 11, "auction": 13,
        "stats": [
            ("Tgt", 5), ("Rec", 6), ("Rec Yds", 7), ("Rec TD", 8), ("PPR", 11),
        ],
    },
    "DST": {
        "sheet": "DST1",
        "player": 2, "team": None, "bye": 3, "fps": 4, "auction": 5,
        "stats": [],
    },
}

JAKES_SHEET = "Jakes Ranks"
COLOR_MAP = {
    "FFFF0000": "ignore", "FF00FF00": "priority", "FFFFFF00": "like",
    "FFFFA500": "caution", "FFFFC000": "caution",
    "FF0000FF": "have", "FF0070C0": "have",
    "FF6AA84F": "rookie", "FF38761D": "rookie",
}
NO_FILL = {None, "00000000", "FFFFFFFF"}


def is_error(v):
    return isinstance(v, str) and v.startswith("#")


def num(v):
    if isinstance(v, (int, float)):
        return round(v, 1)
    return None


def clean_comment(text):
    if not text:
        return None
    lines = text.split("\n")
    body, skipping = [], True
    for line in lines:
        if skipping:
            s = line.strip()
            if s == "======" or s.startswith("ID#"):
                continue
            if re.match(r"^.+\(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}\)\s*$", s):
                continue
            skipping = False
        body.append(line)
    return "\n".join(body).strip() or text.strip()


def load_jakes_tags(wb):
    """Map QB player name -> color tag from the left-hand QB table."""
    tags = {}
    if JAKES_SHEET not in wb.sheetnames:
        return tags
    ws = wb[JAKES_SHEET]
    for row in ws.iter_rows(min_row=2, max_col=14):
        player = row[1].value
        if not player or is_error(player):
            continue
        cell = row[1]
        rgb = cell.fill.fgColor.rgb if (cell.fill and cell.fill.fgColor) else None
        tags[player] = COLOR_MAP.get(rgb) if rgb not in NO_FILL else None
    return tags


def player_for_comment_col(col):
    for lo, hi, pcol in JAKES_TABLE_PLAYER_COL:
        if lo <= col <= hi:
            return pcol
    return None


def load_jakes_comments(wb):
    """Collect every hand-written comment on the sheet -> [(player, ts, body)]."""
    out = []
    if JAKES_SHEET not in wb.sheetnames:
        return out
    ws = wb[JAKES_SHEET]
    for row in ws.iter_rows():
        for cell in row:
            if not cell.comment:
                continue
            pcol = player_for_comment_col(cell.column)
            if not pcol:
                continue
            player = ws.cell(cell.row, pcol).value
            if not player or is_error(player):
                continue
            raw = cell.comment.text
            m = TS_RE.search(raw)
            ts = m.group(1) if m else None
            body = clean_comment(raw)
            if body:
                out.append((player, ts, body))
    return out


def sql_escape(s):
    return s.replace("\\", "\\\\").replace("'", "''")


def write_seed_comments(comments):
    """Idempotent seed: insert each sheet comment as a normal 'Ryan' comment."""
    lines = [
        "-- Auto-generated by export_data.py — Jake's spreadsheet notes as comments.",
        "-- Safe to run once in phpMyAdmin after schema.sql. Re-running skips duplicates.",
        "",
    ]
    for player, ts, body in comments:
        created = ts if ts else "NOW()"
        created_sql = f"'{ts}'" if ts else "NOW()"
        p = sql_escape(player)
        b = sql_escape(body)
        lines.append(
            "INSERT INTO comments (player, author, text, created_at)\n"
            f"SELECT '{p}', 'Ryan', '{b}', {created_sql}\n"
            "WHERE NOT EXISTS (SELECT 1 FROM comments "
            f"WHERE player = '{p}' AND text = '{b}');"
        )
    SEED_PATH.write_text("\n".join(lines) + "\n")
    return len(comments)


def main():
    wb = load_workbook(XLSX_PATH, data_only=True)
    tags = load_jakes_tags(wb)

    all_players = []
    for pos, cfg in POSITIONS.items():
        ws = wb[cfg["sheet"]]
        for r in range(2, 1001):
            name = ws.cell(r, cfg["player"]).value
            if not name or is_error(name):
                continue
            all_players.append({
                "position": pos,
                "player": name,
                "team": ws.cell(r, cfg["team"]).value if cfg["team"] else None,
                "bye": ws.cell(r, cfg["bye"]).value,
                "fps": num(ws.cell(r, cfg["fps"]).value) if cfg["fps"] else None,
                "auction": num(ws.cell(r, cfg["auction"]).value) if cfg["auction"] else None,
                "stats": {label: num(ws.cell(r, col).value) for label, col in cfg["stats"]},
                "tag": tags.get(name),
            })

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(json.dumps({"players": all_players}, indent=2))
    by_pos = {}
    for p in all_players:
        by_pos[p["position"]] = by_pos.get(p["position"], 0) + 1
    print(f"Wrote {len(all_players)} players to {OUTPUT_PATH.relative_to(SCRIPT_DIR)}: {by_pos}")

    n = write_seed_comments(load_jakes_comments(wb))
    print(f"Wrote {n} spreadsheet comments to {SEED_PATH.relative_to(SCRIPT_DIR)}")


if __name__ == "__main__":
    main()
